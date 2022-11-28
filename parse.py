if __name__ == '__main__':

    import openpyxl
    import csv

    # Define variable to load the dataframe
    dataframe = openpyxl.load_workbook("input/Fatura-Excel2.xlsx")

    # Define variable to read sheet
    dataframe1 = dataframe.active

    rowdata = []
    csv_rowlist = []
    # Iterate the loop to read the cell values
    for row in dataframe1.iter_rows(1, dataframe1.max_row):
        if row[0].value == 'data':
            for col in range(0, 3):
                rowdata.append(row[col].value)
                if col == 3:
                    csv_rowlist.append(rowdata)


    with open('output/out.csv', 'w') as file:
        writer = csv.writer(file)
        writer.writerows(csv_rowlist)